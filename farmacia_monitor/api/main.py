"""
API FastAPI — PharmaFlow Backend v2
- Super Admin: acesso total, criado via ADMIN_SECRET do .env
- Gestores: acesso apenas às suas próprias farmácias
"""

import io
import os
import asyncio
from datetime import datetime, timedelta
from typing import Optional

from fastapi import FastAPI, Depends, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from sqlalchemy.orm import Session
from sqlalchemy import text
from jose import JWTError, jwt
import bcrypt as _bcrypt
from pydantic import BaseModel

from farmacia_monitor.database.db import (
    get_db, SessionLocal, GestorTrafego, Farmacia, Coleta
)

# ── Config ────────────────────────────────────────────────────────────────────

SECRET_KEY         = os.getenv("JWT_SECRET_KEY", "troque-no-env-do-servidor")
ALGORITHM          = "HS256"
TOKEN_EXPIRE_HORAS = 8

def _hash_senha(senha: str) -> str:
    return _bcrypt.hashpw(senha.encode(), _bcrypt.gensalt()).decode()

def _verificar_senha(senha: str, hashed: str) -> bool:
    return _bcrypt.checkpw(senha.encode(), hashed.encode())
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/auth/login")

app = FastAPI(title="PharmaFlow API", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ── JWT Helpers ───────────────────────────────────────────────────────────────

def _criar_token(gestor_id: int, nome: str, is_admin: bool) -> str:
    payload = {
        "sub":      str(gestor_id),
        "nome":     nome,
        "is_admin": is_admin,
        "exp":      datetime.utcnow() + timedelta(hours=TOKEN_EXPIRE_HORAS),
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def get_current_user(
    token: str = Depends(oauth2_scheme),
    db: Session = Depends(get_db),
) -> GestorTrafego:
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        gestor_id = payload.get("sub")
        if not gestor_id:
            raise HTTPException(status_code=401, detail="Token invalido")
    except JWTError:
        raise HTTPException(status_code=401, detail="Token invalido ou expirado")

    gestor = db.query(GestorTrafego).filter(
        GestorTrafego.id == int(gestor_id),
        GestorTrafego.ativo == True,
    ).first()
    if not gestor:
        raise HTTPException(status_code=401, detail="Usuario nao encontrado")
    return gestor


def admin_required(current_user: GestorTrafego = Depends(get_current_user)) -> GestorTrafego:
    """Dependência: bloqueia acesso se não for super admin."""
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Acesso restrito ao administrador")
    return current_user


# ── Pydantic Schemas ──────────────────────────────────────────────────────────

class SuperAdminCreate(BaseModel):
    nome:         str
    email:        str
    senha:        str
    admin_secret: str   # deve bater com ADMIN_SECRET do .env

class GestorCreate(BaseModel):
    nome:  str
    email: str
    senha: str

class GestorUpdate(BaseModel):
    nome:  Optional[str] = None
    email: Optional[str] = None
    senha: Optional[str] = None

class FarmaciaCreate(BaseModel):
    nome:      str
    url_base:  str
    email:     str
    senha:     str
    gestor_id: Optional[int] = None

class FarmaciaUpdate(BaseModel):
    nome:         Optional[str]   = None
    url_base:     Optional[str]   = None
    email:        Optional[str]   = None
    senha:        Optional[str]   = None
    gestor_id:    Optional[int]   = None
    ativa:        Optional[bool]  = None
    meta_vendas:  Optional[int]   = None
    meta_receita: Optional[float] = None

class MetaUpdate(BaseModel):
    meta_vendas:  Optional[int]   = None
    meta_receita: Optional[float] = None


# ── Auth ──────────────────────────────────────────────────────────────────────

@app.post("/api/auth/criar-super-admin", status_code=201)
def criar_super_admin(dados: SuperAdminCreate, db: Session = Depends(get_db)):
    """
    Cria o primeiro super admin. Requer ADMIN_SECRET do .env.
    Bloqueado permanentemente após o primeiro admin ser criado.
    """
    secret_env = os.getenv("ADMIN_SECRET", "")
    if not secret_env:
        raise HTTPException(status_code=503, detail="ADMIN_SECRET nao configurado no servidor")
    if dados.admin_secret != secret_env:
        raise HTTPException(status_code=403, detail="Segredo invalido")

    ja_existe = db.query(GestorTrafego).filter(GestorTrafego.is_admin == True).first()
    if ja_existe:
        raise HTTPException(status_code=409, detail="Super admin ja existe. Use o login normal.")

    email_existe = db.query(GestorTrafego).filter(GestorTrafego.email == dados.email).first()
    if email_existe:
        raise HTTPException(status_code=409, detail="Email ja cadastrado")

    admin = GestorTrafego(
        nome=dados.nome,
        email=dados.email,
        senha_hash=_hash_senha(dados.senha),
        is_admin=True,
    )
    db.add(admin)
    db.commit()
    db.refresh(admin)
    return {"id": admin.id, "nome": admin.nome, "email": admin.email, "is_admin": True}


@app.post("/api/auth/login")
def login(form: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    gestor = db.query(GestorTrafego).filter(GestorTrafego.email == form.username).first()
    if not gestor or not _verificar_senha(form.password, gestor.senha_hash):
        raise HTTPException(status_code=401, detail="Email ou senha incorretos")
    if not gestor.ativo:
        raise HTTPException(status_code=403, detail="Usuario inativo")

    return {
        "access_token": _criar_token(gestor.id, gestor.nome, gestor.is_admin),
        "token_type":   "bearer",
        "id":           gestor.id,
        "nome":         gestor.nome,
        "email":        gestor.email,
        "is_admin":     gestor.is_admin,
    }


@app.get("/api/auth/me")
def get_me(current_user: GestorTrafego = Depends(get_current_user)):
    return {
        "id":       current_user.id,
        "nome":     current_user.nome,
        "email":    current_user.email,
        "is_admin": current_user.is_admin,
    }


# ── Gestores CRUD (somente admin) ─────────────────────────────────────────────

@app.get("/api/gestores")
def get_gestores(
    db: Session = Depends(get_db),
    _: GestorTrafego = Depends(get_current_user),   # qualquer logado pode listar
):
    gestores = db.query(GestorTrafego).filter(GestorTrafego.ativo == True).all()
    return [
        {
            "id":        g.id,
            "nome":      g.nome,
            "email":     g.email,
            "is_admin":  g.is_admin,
            "criado_em": g.criado_em,
            "farmacias": len([f for f in g.farmacias if f.ativa]),
        }
        for g in gestores
    ]


@app.post("/api/gestores", status_code=201)
def criar_gestor(
    dados: GestorCreate,
    db: Session = Depends(get_db),
    _: GestorTrafego = Depends(admin_required),   # somente admin
):
    existe = db.query(GestorTrafego).filter(GestorTrafego.email == dados.email).first()
    if existe:
        raise HTTPException(status_code=409, detail="Email ja cadastrado")

    gestor = GestorTrafego(
        nome=dados.nome,
        email=dados.email,
        senha_hash=_hash_senha(dados.senha),
        is_admin=False,
    )
    db.add(gestor)
    db.commit()
    db.refresh(gestor)
    return {"id": gestor.id, "nome": gestor.nome, "email": gestor.email}


@app.put("/api/gestores/{gestor_id}")
def atualizar_gestor(
    gestor_id: int,
    dados: GestorUpdate,
    db: Session = Depends(get_db),
    _: GestorTrafego = Depends(admin_required),
):
    gestor = db.query(GestorTrafego).filter(GestorTrafego.id == gestor_id).first()
    if not gestor:
        raise HTTPException(status_code=404, detail="Gestor nao encontrado")

    if dados.nome  is not None: gestor.nome  = dados.nome
    if dados.email is not None: gestor.email = dados.email
    if dados.senha:             gestor.senha_hash = _hash_senha(dados.senha)

    db.commit()
    return {"id": gestor.id, "nome": gestor.nome, "email": gestor.email}


@app.delete("/api/gestores/{gestor_id}")
def deletar_gestor(
    gestor_id: int,
    db: Session = Depends(get_db),
    current_user: GestorTrafego = Depends(admin_required),
):
    if gestor_id == current_user.id:
        raise HTTPException(status_code=400, detail="Nao e possivel deletar seu proprio usuario")

    gestor = db.query(GestorTrafego).filter(GestorTrafego.id == gestor_id).first()
    if not gestor:
        raise HTTPException(status_code=404, detail="Gestor nao encontrado")

    gestor.ativo = False
    db.commit()
    return {"mensagem": "Gestor desativado"}


# ── Farmácias CRUD ────────────────────────────────────────────────────────────

@app.post("/api/farmacias", status_code=201)
def criar_farmacia(
    dados: FarmaciaCreate,
    db: Session = Depends(get_db),
    _: GestorTrafego = Depends(admin_required),
):
    from farmacia_monitor.cripto import _fernet
    farmacia = Farmacia(
        nome=dados.nome,
        url_base=dados.url_base,
        email=dados.email,
        senha_enc=_fernet().encrypt(dados.senha.encode()).decode(),
        gestor_id=dados.gestor_id,
    )
    db.add(farmacia)
    db.commit()
    db.refresh(farmacia)
    return {"id": farmacia.id, "nome": farmacia.nome, "gestor_id": farmacia.gestor_id}


@app.put("/api/farmacias/{farmacia_id}")
def atualizar_farmacia(
    farmacia_id: int,
    dados: FarmaciaUpdate,
    db: Session = Depends(get_db),
    _: GestorTrafego = Depends(admin_required),
):
    farmacia = db.query(Farmacia).filter(Farmacia.id == farmacia_id).first()
    if not farmacia:
        raise HTTPException(status_code=404, detail="Farmacia nao encontrada")

    if dados.nome         is not None: farmacia.nome         = dados.nome
    if dados.url_base     is not None: farmacia.url_base     = dados.url_base
    if dados.email        is not None: farmacia.email        = dados.email
    if dados.gestor_id    is not None: farmacia.gestor_id    = dados.gestor_id
    if dados.ativa        is not None: farmacia.ativa        = dados.ativa
    if dados.meta_vendas  is not None: farmacia.meta_vendas  = dados.meta_vendas
    if dados.meta_receita is not None: farmacia.meta_receita = dados.meta_receita
    if dados.senha:
        from farmacia_monitor.cripto import _fernet
        farmacia.senha_enc = _fernet().encrypt(dados.senha.encode()).decode()

    db.commit()
    return {"id": farmacia.id, "nome": farmacia.nome, "ativa": farmacia.ativa}


@app.patch("/api/farmacias/{farmacia_id}/meta")
def atualizar_meta(
    farmacia_id: int,
    dados: MetaUpdate,
    db: Session = Depends(get_db),
    _: GestorTrafego = Depends(admin_required),
):
    """Define ou atualiza a meta semanal de vendas e/ou receita de uma farmácia."""
    farmacia = db.query(Farmacia).filter(Farmacia.id == farmacia_id).first()
    if not farmacia:
        raise HTTPException(status_code=404, detail="Farmacia nao encontrada")

    if dados.meta_vendas  is not None: farmacia.meta_vendas  = dados.meta_vendas
    if dados.meta_receita is not None: farmacia.meta_receita = dados.meta_receita

    db.commit()
    return {
        "id":           farmacia.id,
        "nome":         farmacia.nome,
        "meta_vendas":  farmacia.meta_vendas,
        "meta_receita": float(farmacia.meta_receita) if farmacia.meta_receita else None,
    }


@app.delete("/api/farmacias/{farmacia_id}")
def deletar_farmacia(
    farmacia_id: int,
    db: Session = Depends(get_db),
    _: GestorTrafego = Depends(admin_required),
):
    farmacia = db.query(Farmacia).filter(Farmacia.id == farmacia_id).first()
    if not farmacia:
        raise HTTPException(status_code=404, detail="Farmacia nao encontrada")

    farmacia.ativa = False
    db.commit()
    return {"mensagem": "Farmacia desativada"}


# ── Painel Geral ──────────────────────────────────────────────────────────────

def _mapear_nome_canal(nome: str) -> str:
    n = nome.lower()
    if "google" in n:
        return "Google"
    if "facebook" in n or "instagram" in n or "meta" in n:
        return "Meta"
    if "grupo" in n or "oferta" in n or "group" in n:
        return "Grupos"
    return nome  # mantém o nome original para canais não mapeados


@app.get("/api/painel")
def get_painel(
    gestor_id: Optional[int] = None,
    current_user: GestorTrafego = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    filtro = gestor_id if current_user.is_admin else current_user.id

    params: dict = {}
    filtro_sql = ""
    if filtro:
        filtro_sql = "AND f.gestor_id = :gid"
        params["gid"] = filtro

    rows = db.execute(text(f"""
        SELECT r.* FROM vw_ranking_atual r
        JOIN farmacias f ON f.id = r.farmacia_id
        WHERE TRUE {filtro_sql}
    """), params).mappings().all()

    if not rows:
        return {
            "receita_total": 0, "total_atendimentos": 0, "vendas_realizadas": 0,
            "farmacias_ativas": 0, "farmacias_alerta": 0, "farmacias_atencao": 0,
            "taxa_conversao_media": 0, "ultima_atualizacao": None, "canais": [],
        }

    receita_total      = sum(float(r["receita_total"] or 0) for r in rows)
    total_atendimentos = sum(int(r["total_atendimentos"] or 0) for r in rows)
    vendas_realizadas  = sum(int(r["vendas_realizadas"] or 0) for r in rows)
    conversoes = [
        round(float(r["vendas_realizadas"] or 0) / float(r["total_atendimentos"]) * 100, 2)
        for r in rows if float(r["total_atendimentos"] or 0) > 0
    ]
    taxa_media = round(sum(conversoes) / len(conversoes), 2) if conversoes else 0
    ultima_atualizacao = max((r["data_coleta"] for r in rows), default=None)

    # Canais: agrega da última coleta de cada farmácia filtrada
    canais_rows = db.execute(text(f"""
        SELECT cc.canal,
               SUM(cc.atendimentos)::int          AS total_atendimentos,
               SUM(cc.vendas)::int                AS total_vendas,
               SUM(cc.receita_vendas)::numeric    AS total_receita_vendas
        FROM coleta_canais cc
        JOIN (
            SELECT DISTINCT ON (farmacia_id) id AS coleta_id, farmacia_id
            FROM coletas
            ORDER BY farmacia_id, data_coleta DESC
        ) latest ON latest.coleta_id = cc.coleta_id
        JOIN farmacias f ON f.id = latest.farmacia_id
        WHERE f.ativa = TRUE {filtro_sql}
        GROUP BY cc.canal
        ORDER BY total_atendimentos DESC
    """), params).mappings().all()

    # Agrega por nome padronizado (Google / Meta / Grupos / outros)
    canais_agg: dict[str, dict] = {}
    for row in canais_rows:
        nome_std = _mapear_nome_canal(row["canal"])
        if nome_std not in canais_agg:
            canais_agg[nome_std] = {"atendimentos": 0, "vendas": 0, "receita_vendas": 0.0}
        canais_agg[nome_std]["atendimentos"]   += int(row["total_atendimentos"] or 0)
        canais_agg[nome_std]["vendas"]         += int(row["total_vendas"] or 0)
        canais_agg[nome_std]["receita_vendas"] += float(row["total_receita_vendas"] or 0)

    canais = [
        {
            "nome":           nome,
            "atendimentos":   dados["atendimentos"],
            "vendas":         dados["vendas"],
            "receita_vendas": round(dados["receita_vendas"], 2),
        }
        for nome, dados in sorted(canais_agg.items(), key=lambda x: -x[1]["atendimentos"])
    ]

    return {
        "receita_total":        round(receita_total, 2),
        "total_atendimentos":   total_atendimentos,
        "vendas_realizadas":    vendas_realizadas,
        "farmacias_ativas":     len(rows),
        "farmacias_alerta":     sum(1 for r in rows if r["nivel_alerta"] == "vermelho"),
        "farmacias_atencao":    sum(1 for r in rows if r["nivel_alerta"] == "amarelo"),
        "taxa_conversao_media": taxa_media,
        "ultima_atualizacao":   ultima_atualizacao,
        "canais":               canais,
    }


# ── Farmácias (listagem + evolução) ───────────────────────────────────────────

@app.get("/api/farmacias")
def get_farmacias(
    status:    Optional[str] = None,
    busca:     Optional[str] = None,
    gestor_id: Optional[int] = None,
    current_user: GestorTrafego = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    # Gestor comum: forçar filtro pelo seu próprio ID
    filtro = gestor_id if current_user.is_admin else current_user.id

    # LEFT JOIN para incluir farmácias sem coletas ainda
    sql = """
        SELECT
            f.id AS farmacia_id, f.nome AS farmacia, f.gestor_id, f.ativa,
            f.meta_vendas, f.meta_receita,
            COALESCE(r.nivel_alerta, 'verde')        AS nivel_alerta,
            COALESCE(r.receita_total, 0)             AS receita_total,
            COALESCE(r.total_atendimentos, 0)        AS total_atendimentos,
            COALESCE(r.vendas_realizadas, 0)         AS vendas_realizadas,
            COALESCE(r.variacao_receita, 0)          AS variacao_receita,
            COALESCE(r.variacao_vendas, 0)           AS variacao_vendas,
            COALESCE(r.score_criticidade, 0)         AS score_criticidade,
            COALESCE(r.posicao_ranking, 9999)        AS posicao_ranking,
            r.periodo_inicio, r.periodo_fim, r.data_coleta
        FROM farmacias f
        LEFT JOIN vw_ranking_atual r ON r.farmacia_id = f.id
        WHERE f.ativa = TRUE
        {filtro_sql}
        ORDER BY posicao_ranking
    """
    if filtro:
        rows = db.execute(
            text(sql.format(filtro_sql="AND f.gestor_id = :gid")), {"gid": filtro}
        ).mappings().all()
    else:
        rows = db.execute(text(sql.format(filtro_sql=""))).mappings().all()

    # Busca canais agregados por farmácia (última coleta de cada uma)
    canais_filtro_sql = "AND f.gestor_id = :gid" if filtro else ""
    canais_params: dict = {}
    if filtro:
        canais_params["gid"] = filtro

    canais_rows = db.execute(text(f"""
        SELECT latest.farmacia_id,
               cc.canal,
               SUM(cc.atendimentos)::int       AS total_atendimentos,
               SUM(cc.vendas)::int             AS total_vendas,
               SUM(cc.receita_vendas)::numeric AS total_receita_vendas
        FROM coleta_canais cc
        JOIN (
            SELECT DISTINCT ON (farmacia_id) id AS coleta_id, farmacia_id
            FROM coletas
            ORDER BY farmacia_id, data_coleta DESC
        ) latest ON latest.coleta_id = cc.coleta_id
        JOIN farmacias f ON f.id = latest.farmacia_id
        WHERE f.ativa = TRUE {canais_filtro_sql}
        GROUP BY latest.farmacia_id, cc.canal
        ORDER BY latest.farmacia_id, total_atendimentos DESC
    """), canais_params).mappings().all()

    # Agrupa canais por farmacia_id com nome padronizado
    canais_por_farmacia: dict[int, dict[str, dict]] = {}
    for cr in canais_rows:
        fid = int(cr["farmacia_id"])
        nome_std = _mapear_nome_canal(cr["canal"])
        canais_por_farmacia.setdefault(fid, {})
        if nome_std not in canais_por_farmacia[fid]:
            canais_por_farmacia[fid][nome_std] = {"atendimentos": 0, "vendas": 0, "receita_vendas": 0.0}
        canais_por_farmacia[fid][nome_std]["atendimentos"]   += int(cr["total_atendimentos"] or 0)
        canais_por_farmacia[fid][nome_std]["vendas"]         += int(cr["total_vendas"] or 0)
        canais_por_farmacia[fid][nome_std]["receita_vendas"] += float(cr["total_receita_vendas"] or 0)

    resultado = []
    for r in rows:
        label_status = {"verde": "Ativa", "amarelo": "Atencao", "vermelho": "Alerta"}.get(
            r["nivel_alerta"], "Ativa"
        )
        if status and label_status.lower() != status.lower():
            continue
        if busca and busca.lower() not in r["farmacia"].lower():
            continue

        fid = int(r["farmacia_id"])
        canais = [
            {
                "nome":           nome,
                "atendimentos":   dados["atendimentos"],
                "vendas":         dados["vendas"],
                "receita_vendas": round(dados["receita_vendas"], 2),
            }
            for nome, dados in sorted(
                canais_por_farmacia.get(fid, {}).items(), key=lambda x: -x[1]["atendimentos"]
            )
        ]

        receita_atual = float(r["receita_total"] or 0)
        vendas_atual  = int(r["vendas_realizadas"] or 0)
        meta_v = r["meta_vendas"]
        meta_r = float(r["meta_receita"] or 0) if r["meta_receita"] else None

        # Calcula se atingiu a meta e percentual de execução
        atingiu_meta = None          # null = sem meta definida
        percentual_meta_receita = 0.0
        percentual_meta_vendas  = 0.0
        if meta_r:
            percentual_meta_receita = round(receita_atual / meta_r * 100, 1)
            atingiu_meta = receita_atual >= meta_r
        if meta_v:
            percentual_meta_vendas = round(vendas_atual / meta_v * 100, 1)
            if atingiu_meta is None:
                atingiu_meta = vendas_atual >= meta_v
            else:
                atingiu_meta = atingiu_meta and (vendas_atual >= meta_v)

        resultado.append({
            "id":                       fid,
            "nome":                     r["farmacia"],
            "status":                   label_status,
            "nivel_alerta":             r["nivel_alerta"],
            "gestor_id":                r.get("gestor_id"),
            "receita_total":            receita_atual,
            "total_atendimentos":       int(r["total_atendimentos"] or 0),
            "atendimentos_finalizados": 0,
            "vendas_realizadas":        vendas_atual,
            "taxa_conversao":           0,
            "variacao_receita":         float(r.get("variacao_receita") or 0),
            "variacao_atendimentos":    0,
            "variacao_vendas":          float(r.get("variacao_vendas") or 0),
            "score_criticidade":        float(r["score_criticidade"] or 0),
            "posicao_ranking":          int(r["posicao_ranking"]),
            "periodo_inicio":           str(r["periodo_inicio"]) if r["periodo_inicio"] else None,
            "periodo_fim":              str(r["periodo_fim"]) if r["periodo_fim"] else None,
            "data_coleta":              r["data_coleta"],
            "meta_vendas":              meta_v,
            "meta_receita":             meta_r,
            "atingiu_meta":             atingiu_meta,
            "percentual_meta_receita":  percentual_meta_receita,
            "percentual_meta_vendas":   percentual_meta_vendas,
            "canais":                   canais,
        })

    return resultado


@app.get("/api/farmacias/{farmacia_id}/evolucao")
def get_evolucao(
    farmacia_id: int,
    current_user: GestorTrafego = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    # Gestor só pode ver evolução das suas próprias farmácias
    if not current_user.is_admin:
        farmacia = db.query(Farmacia).filter(
            Farmacia.id == farmacia_id,
            Farmacia.gestor_id == current_user.id,
        ).first()
        if not farmacia:
            raise HTTPException(status_code=403, detail="Acesso negado a esta farmacia")

    rows = db.execute(text("""
        SELECT * FROM vw_evolucao_semanal
        WHERE farmacia_id = :fid
        ORDER BY semana_numero ASC
    """), {"fid": farmacia_id}).mappings().all()
    return [dict(r) for r in rows]


# ── Relatórios ────────────────────────────────────────────────────────────────

@app.get("/api/relatorios")
def get_relatorios(
    _: GestorTrafego = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    rows = db.execute(text("""
        SELECT
            DATE_TRUNC('week', data_coleta)::DATE  AS periodo_inicio,
            MAX(periodo_fim)                        AS periodo_fim,
            MAX(data_coleta)                        AS data_geracao,
            COUNT(DISTINCT farmacia_id)             AS farmacias,
            SUM(CASE WHEN nivel_alerta != 'sem_dados' THEN 1 ELSE 0 END) AS concluidas
        FROM coletas
        GROUP BY DATE_TRUNC('week', data_coleta)::DATE
        ORDER BY periodo_inicio DESC
        LIMIT 20
    """)).mappings().all()

    resultado = []
    for i, r in enumerate(rows):
        total      = int(r["farmacias"] or 0)
        concluidas = int(r["concluidas"] or total)
        status     = "Concluido" if concluidas == total else "Parcial" if concluidas > 0 else "Erro"
        inicio     = r["periodo_inicio"]
        fim        = r["periodo_fim"]
        resultado.append({
            "id":             i + 1,
            "label":          f"Semana {len(rows) - i} — {_fmt_data(inicio)} a {_fmt_data(fim)}",
            "periodo_inicio": str(inicio),
            "periodo_fim":    str(fim),
            "data_geracao":   r["data_geracao"],
            "farmacias":      f"{concluidas}/{total}",
            "status":         status,
        })
    return resultado


def _query_relatorio(db: Session, periodo_inicio: str):
    """Query base compartilhada pelos downloads XLSX e CSV."""
    rows = db.execute(text("""
        SELECT
            f.nome                              AS farmacia,
            g.nome                              AS gestor,
            c.periodo_inicio, c.periodo_fim,
            c.receita_total, c.total_atendimentos,
            c.vendas_realizadas, c.score_criticidade, c.nivel_alerta,
            f.meta_receita, f.meta_vendas,
            CASE
                WHEN f.meta_receita IS NOT NULL AND c.receita_total < f.meta_receita THEN 'Nao'
                WHEN f.meta_vendas  IS NOT NULL AND c.vendas_realizadas < f.meta_vendas THEN 'Nao'
                WHEN f.meta_receita IS NULL AND f.meta_vendas IS NULL THEN 'Sem meta'
                ELSE 'Sim'
            END AS atingiu_meta,
            CASE WHEN f.meta_receita > 0
                 THEN ROUND(c.receita_total / f.meta_receita * 100, 1)
                 ELSE NULL END AS pct_meta_receita
        FROM coletas c
        JOIN farmacias f ON f.id = c.farmacia_id
        LEFT JOIN gestores_trafego g ON g.id = f.gestor_id
        WHERE c.periodo_inicio::TEXT = :periodo
        ORDER BY c.score_criticidade DESC
    """), {"periodo": periodo_inicio}).mappings().all()
    return rows


def _query_canais_relatorio(db: Session, periodo_inicio: str):
    """Canais de vendas para o período (uma linha por farmácia+canal)."""
    return db.execute(text("""
        SELECT
            f.nome AS farmacia,
            cc.canal,
            cc.atendimentos,
            cc.vendas,
            cc.receita_vendas
        FROM coleta_canais cc
        JOIN coletas c  ON c.id  = cc.coleta_id
        JOIN farmacias f ON f.id = c.farmacia_id
        WHERE c.periodo_inicio::TEXT = :periodo
        ORDER BY f.nome, cc.atendimentos DESC
    """), {"periodo": periodo_inicio}).mappings().all()


@app.get("/api/relatorios/{periodo_inicio}/xlsx")
def download_xlsx(
    periodo_inicio: str,
    _: GestorTrafego = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, numbers

    rows       = _query_relatorio(db, periodo_inicio)
    canal_rows = _query_canais_relatorio(db, periodo_inicio)

    if not rows:
        raise HTTPException(status_code=404, detail="Periodo nao encontrado")

    wb = openpyxl.Workbook()

    # ── Aba 1: Resumo por farmácia ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumo"

    cabecalho = [
        "Farmácia", "Gestor", "Período Início", "Período Fim",
        "Receita (R$)", "Meta Receita (R$)", "% Meta Receita",
        "Vendas", "Meta Vendas", "Atingiu Meta",
        "Atendimentos", "Score", "Alerta",
    ]
    header_fill = PatternFill("solid", fgColor="1A7A4A")
    header_font = Font(bold=True, color="FFFFFF")
    for col, titulo in enumerate(cabecalho, 1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(titulo) + 2, 16)

    cores = {"verde": "C6EFCE", "amarelo": "FFEB9C", "vermelho": "FFC7CE"}
    for linha, r in enumerate(rows, 2):
        pct = float(r["pct_meta_receita"]) / 100 if r["pct_meta_receita"] else None
        valores = [
            r["farmacia"],
            r["gestor"] or "—",
            str(r["periodo_inicio"]),
            str(r["periodo_fim"]),
            float(r["receita_total"] or 0),
            float(r["meta_receita"]) if r["meta_receita"] else "—",
            pct,
            int(r["vendas_realizadas"] or 0),
            int(r["meta_vendas"]) if r["meta_vendas"] else "—",
            r["atingiu_meta"],
            int(r["total_atendimentos"] or 0),
            float(r["score_criticidade"] or 0),
            r["nivel_alerta"],
        ]
        fill = PatternFill("solid", fgColor=cores.get(r["nivel_alerta"], "FFFFFF"))
        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=linha, column=col, value=val)
            cell.fill = fill
            if col == 7 and pct is not None:   # % meta
                cell.number_format = "0.0%"
            if col in (5, 6):                  # receita
                cell.number_format = '#,##0.00'

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Aba 2: Canais de vendas ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Canais de Vendas")
    cab2 = ["Farmácia", "Canal", "Atendimentos", "Vendas", "Receita (R$)"]
    for col, titulo in enumerate(cab2, 1):
        cell = ws2.cell(row=1, column=col, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws2.column_dimensions[cell.column_letter].width = max(len(titulo) + 4, 18)

    for linha, r in enumerate(canal_rows, 2):
        ws2.cell(row=linha, column=1, value=r["farmacia"])
        ws2.cell(row=linha, column=2, value=r["canal"])
        ws2.cell(row=linha, column=3, value=int(r["atendimentos"] or 0))
        ws2.cell(row=linha, column=4, value=int(r["vendas"] or 0))
        cell = ws2.cell(row=linha, column=5, value=float(r["receita_vendas"] or 0))
        cell.number_format = '#,##0.00'

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=relatorio_{periodo_inicio}.xlsx"},
    )


@app.get("/api/relatorios/{periodo_inicio}/csv")
def download_csv(
    periodo_inicio: str,
    _: GestorTrafego = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """CSV otimizado para Power BI — UTF-8 BOM, separador vírgula, dados completos."""
    import csv

    rows       = _query_relatorio(db, periodo_inicio)
    canal_rows = _query_canais_relatorio(db, periodo_inicio)

    if not rows:
        raise HTTPException(status_code=404, detail="Periodo nao encontrado")

    buffer = io.StringIO()
    buffer.write("﻿")   # BOM UTF-8 para Power BI
    writer = csv.writer(buffer)

    # Tabela principal
    writer.writerow([
        "Farmacia", "Gestor", "Periodo_Inicio", "Periodo_Fim",
        "Receita_BRL", "Meta_Receita_BRL", "Pct_Meta_Receita",
        "Vendas_Realizadas", "Meta_Vendas", "Atingiu_Meta",
        "Total_Atendimentos", "Score_Criticidade", "Nivel_Alerta",
    ])
    for r in rows:
        writer.writerow([
            r["farmacia"],
            r["gestor"] or "",
            str(r["periodo_inicio"]),
            str(r["periodo_fim"]),
            float(r["receita_total"] or 0),
            float(r["meta_receita"]) if r["meta_receita"] else "",
            float(r["pct_meta_receita"]) if r["pct_meta_receita"] else "",
            int(r["vendas_realizadas"] or 0),
            int(r["meta_vendas"]) if r["meta_vendas"] else "",
            r["atingiu_meta"],
            int(r["total_atendimentos"] or 0),
            float(r["score_criticidade"] or 0),
            r["nivel_alerta"],
        ])

    # Linha em branco + tabela de canais
    writer.writerow([])
    writer.writerow(["--- CANAIS DE VENDAS ---"])
    writer.writerow(["Farmacia", "Canal", "Atendimentos", "Vendas", "Receita_BRL"])
    for r in canal_rows:
        writer.writerow([
            r["farmacia"], r["canal"],
            int(r["atendimentos"] or 0),
            int(r["vendas"] or 0),
            float(r["receita_vendas"] or 0),
        ])

    content = buffer.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        io.BytesIO(content),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=relatorio_{periodo_inicio}.csv"},
    )


# ── Pipeline manual ───────────────────────────────────────────────────────────

_pipeline_rodando = False

@app.post("/api/rodar-agora")
async def rodar_agora(
    background_tasks: BackgroundTasks,
    _: GestorTrafego = Depends(admin_required),
):
    global _pipeline_rodando
    if _pipeline_rodando:
        return {"status": "ja_rodando", "mensagem": "Pipeline ja esta em execucao"}

    async def _executar():
        global _pipeline_rodando
        _pipeline_rodando = True
        try:
            from main import pipeline
            await pipeline()
        finally:
            _pipeline_rodando = False

    background_tasks.add_task(_executar)
    return {"status": "iniciado", "mensagem": "Pipeline iniciado em background"}


@app.get("/api/status")
def get_status():
    return {"pipeline_rodando": _pipeline_rodando, "timestamp": datetime.utcnow().isoformat()}


# ── Ranking de Gestores ───────────────────────────────────────────────────────

@app.get("/api/ranking/gestores")
def get_ranking_gestores(
    current_user: GestorTrafego = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Retorna o ranking dos gestores de tráfego ordenado por % de meta atingida.
    Gestores sem meta definida nas farmácias usam variação de receita como proxy.
    """
    gestores = db.query(GestorTrafego).filter(GestorTrafego.ativo == True).all()

    ranking = []
    for g in gestores:
        farmacias_ativas = [f for f in g.farmacias if f.ativa]
        if not farmacias_ativas:
            continue

        total_farmacias     = len(farmacias_ativas)
        farmacias_com_meta  = 0
        farmacias_ok        = 0
        soma_pct_meta       = 0.0
        receita_total       = 0.0
        meta_receita_total  = 0.0
        vendas_total        = 0
        meta_vendas_total   = 0

        for f in farmacias_ativas:
            # Pega a coleta mais recente
            coleta = (
                db.query(Coleta)
                .filter(Coleta.farmacia_id == f.id)
                .order_by(Coleta.data_coleta.desc())
                .first()
            )
            if not coleta:
                continue

            receita_atual = float(coleta.receita_total or 0)
            vendas_atual  = int(coleta.vendas_realizadas or 0)
            receita_total += receita_atual
            vendas_total  += vendas_atual

            meta_r = float(f.meta_receita or 0)
            meta_v = int(f.meta_vendas or 0)

            if meta_r or meta_v:
                farmacias_com_meta += 1
                meta_receita_total += meta_r
                meta_vendas_total  += meta_v

                # Percentual de meta baseado em receita (principal) ou vendas
                if meta_r:
                    pct = min(receita_atual / meta_r * 100, 150)  # cap 150%
                elif meta_v:
                    pct = min(vendas_atual / meta_v * 100, 150)
                else:
                    pct = 100.0

                soma_pct_meta += pct
                if pct >= 100:
                    farmacias_ok += 1

        # Pontos: 1 ponto por farmácia que bateu a meta
        pontos = farmacias_ok

        if farmacias_com_meta > 0:
            percentual_medio = round(soma_pct_meta / farmacias_com_meta, 1)
            taxa_acerto = round(farmacias_ok / farmacias_com_meta * 100, 1)
        else:
            percentual_medio = 0.0
            taxa_acerto      = 0.0

        ranking.append({
            "gestor_id":             g.id,
            "gestor_nome":           g.nome,
            "total_farmacias":       total_farmacias,
            "farmacias_com_meta":    farmacias_com_meta,
            "farmacias_meta_ok":     farmacias_ok,
            "pontos":                pontos,          # 1 ponto por farmácia que bateu a meta
            "taxa_acerto":           taxa_acerto,     # % das farmácias com meta que bateram
            "percentual_medio_meta": percentual_medio, # % médio atingido da meta
            "tem_meta":              farmacias_com_meta > 0,
            "receita_total":         round(receita_total, 2),   # só informativo
            "vendas_total":          vendas_total,              # só informativo
        })

    # Ordena por pontos (farmácias que bateram a meta) — NÃO por faturamento
    ranking.sort(key=lambda x: (-x["pontos"], -x["taxa_acerto"]))
    for i, item in enumerate(ranking, 1):
        item["posicao"] = i

    return ranking


# ── Helper ────────────────────────────────────────────────────────────────────

def _fmt_data(d) -> str:
    if not d: return ""
    if hasattr(d, "strftime"): return d.strftime("%d %b")
    return str(d)
